from openpyxl import load_workbook
from pathlib import Path
from urllib.parse import quote
from difflib import SequenceMatcher
from html import unescape
import os, re, json, requests

WORKBOOK = Path('/home/ubuntu/upload/ملفالصناديق.xlsx')
BASE_URL = os.environ.get('SUPABASE_URL')
SECRET = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
if not BASE_URL or not SECRET:
    raise RuntimeError('Supabase server configuration is missing')
HEADERS = {'apikey': SECRET, 'Authorization': f'Bearer {SECRET}'}

def normalize(value):
    value = unescape(str(value or ''))
    value = value.replace('\u2013', '-').replace('\u2014', '-').replace('\u2011', '-')
    value = value.casefold()
    value = re.sub(r'[()\[\]{}.,:/\\"\'“”‘’]', ' ', value)
    value = re.sub(r'[‐‑‒–—―-]', ' ', value)
    value = re.sub(r'\s+', ' ', value).strip()
    return value

def get_all(table, select, limit=1000):
    url = f'{BASE_URL}/rest/v1/{table}?select={quote(select, safe=",")}\u0026limit={limit}'
    response = requests.get(url, headers=HEADERS, timeout=60)
    if not response.ok:
        raise RuntimeError(f'{table} lookup failed: {response.status_code} {response.text[:300]}')
    return response.json()

wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header = [str(v or '').strip() for v in rows[0]]
fund_col = header.index('Fund')
company_col = header.index('Management Company')
file_funds = []
for row_number, row in enumerate(rows[1:], start=2):
    name = str(row[fund_col] or '').strip()
    if name:
        file_funds.append({'row': row_number, 'name': name, 'company': str(row[company_col] or '').strip()})

db_funds = get_all('funds', 'fund_id,canonical_name,eima_name_raw,management_company_raw,price_update_url,source_id,active', 1000)
db_prices = get_all('fund_prices', 'fund_id,nav,currency,valuation_date,status,source_id,collected_at', 10000)
prices_by_fund = {}
for price in db_prices:
    if price.get('status') not in (None, 'validated'):
        continue
    fund_id = price.get('fund_id')
    if not fund_id:
        continue
    key = (price.get('valuation_date') or '', price.get('collected_at') or '')
    if key > prices_by_fund.get(fund_id, ('', ''))[0:2]:
        prices_by_fund[fund_id] = (key[0], key[1], price)

name_index = {}
for fund in db_funds:
    for raw in (fund.get('canonical_name'), fund.get('eima_name_raw')):
        key = normalize(raw)
        if key:
            name_index.setdefault(key, []).append(fund)

matches = []
for item in file_funds:
    key = normalize(item['name'])
    candidates = name_index.get(key, [])
    match_type = 'exact' if candidates else None
    if not candidates:
        # Conservative fallback: only accept a unique high-similarity candidate.
        scored = []
        for fund in db_funds:
            for raw in (fund.get('canonical_name'), fund.get('eima_name_raw')):
                candidate_key = normalize(raw)
                if not candidate_key:
                    continue
                score = SequenceMatcher(None, key, candidate_key).ratio()
                scored.append((score, fund))
        scored.sort(key=lambda pair: pair[0], reverse=True)
        if scored and scored[0][0] >= 0.93 and (len(scored) == 1 or scored[0][0] - scored[1][0] >= 0.03):
            candidates = [scored[0][1]]
            match_type = 'high_similarity'
    unique = {fund['fund_id']: fund for fund in candidates}
    if len(unique) == 1:
        fund = next(iter(unique.values()))
        latest = prices_by_fund.get(fund['fund_id'])
        matches.append({
            'row': item['row'], 'file_name': item['name'], 'company': item['company'],
            'match_type': match_type, 'fund_id': fund['fund_id'],
            'db_name': fund.get('canonical_name'), 'db_raw_name': fund.get('eima_name_raw'),
            'db_company': fund.get('management_company_raw'), 'price_update_url': fund.get('price_update_url'),
            'has_price': bool(latest), 'latest_price': latest[2] if latest else None,
        })
    else:
        matches.append({'row': item['row'], 'file_name': item['name'], 'company': item['company'], 'match_type': 'ambiguous' if unique else 'unmatched', 'candidates': [f.get('canonical_name') for f in unique.values()]})

covered = [m for m in matches if m.get('has_price')]
in_db = [m for m in matches if m.get('fund_id')]
uncovered = [m for m in matches if not m.get('has_price')]
report = {
    'workbook': str(WORKBOOK), 'sheet': ws.title, 'file_fund_count': len(file_funds),
    'database_fund_count': len(db_funds), 'database_price_row_count': len(db_prices),
    'matched_to_database_count': len(in_db), 'covered_with_validated_price_count': len(covered),
    'not_covered_count': len(uncovered), 'exact_or_confident_match_count': len([m for m in in_db if m.get('match_type') in ('exact','high_similarity')]),
    'unmatched_or_ambiguous_count': len([m for m in matches if not m.get('fund_id')]),
    'covered': covered, 'not_covered': uncovered,
}
Path('/tmp/upload-coverage-report.json').write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps({k: report[k] for k in ('file_fund_count','database_fund_count','database_price_row_count','matched_to_database_count','covered_with_validated_price_count','not_covered_count','unmatched_or_ambiguous_count')}, ensure_ascii=False, indent=2))
print('--- NOT COVERED ---')
for item in uncovered:
    print(f"{item['row']}\t{item['file_name']}\t{item.get('match_type')}\t{'; '.join(item.get('candidates', []))}")
