from openpyxl import load_workbook
from pathlib import Path
import json

path = Path('/home/ubuntu/upload/ملفالصناديق.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
result = {}
for ws in wb.worksheets:
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = [None if value is None else str(value).strip() for value in row]
        if any(value not in (None, '') for value in values):
            rows.append(values)
    result[ws.title] = {'max_row': ws.max_row, 'max_column': ws.max_column, 'rows': rows}
print(json.dumps(result, ensure_ascii=False, indent=2))
